from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Student_data, Faculty_profile, HOD_profile, StudentFacultyMapping
from django.http import HttpResponse
import openpyxl

def main_dashboard(request):
    return render(request, "main_dashboard.html")

def Student_Login(request):
    if request.method == "POST":
        email = request.POST.get("email")  # Change 'username' to 'email'
        password = request.POST.get("password")

        try:
            # Check if the student exists in the Student_data model
            student = Student_data.objects.get(email=email, password=password)

            # Check if there is a mapping for this student in StudentFacultyMapping
            try:
                mapping = StudentFacultyMapping.objects.get(student_email=student.email)  # Change 'student_name' to 'student_email'

                # If a mapping exists, retrieve faculty information
                faculty_name = mapping.faculty_name
                faculty_email = mapping.faculty_email
                faculty_phone = mapping.faculty_phone
                student_name = mapping.student_name


                # Render the student dashboard with faculty information
                return render(request, "student_dashboard.html", {
                    "email": email,  # Update the variable name to 'email'
                    "faculty_name": faculty_name,
                    "faculty_email": faculty_email,
                    "faculty_phone": faculty_phone,
                    "student_name": student_name,
                })
            except StudentFacultyMapping.DoesNotExist:
                # Handle the case where there is no mapping for this student
                error_message = "Student has no faculty mapping"
                return render(request, "Student_Login.html", {"error_message": error_message})

        except Student_data.DoesNotExist:
            messages.error(request,"Invalid email or password")  # Update the error message
            return render(request, "Student_Login.html")

    return render(request, "Student_Login.html")


# def student_dashboard(request):
#     # Redirect to login page if not authenticated
#     if not request.user.is_authenticated:
#         return render(request, "Student_Login.html")
#
#     # If authenticated, display the student dashboard
#     return render(request, "student_dashboard.html")
from .models import Circular

from .models import Circular


def student_dashboard(request):
    # Redirect to login page if not authenticated
    if not request.user.is_authenticated:
        return render(request, "Student_Login.html")

    # Retrieve circulars from the database
    circulars = Circular.objects.all()

    print(circulars)  # Debugging statement

    # Pass circulars to the template context
    context = {'circulars': circulars}
    return render(request, "student_dashboard.html", context)

def display_circulars(request):
    # Fetch all circulars from the database
    all_circulars = Circular.objects.all()

    # Pass circulars to the template context
    context = {'all_circulars': all_circulars}
    return render(request, "dashboard.html", context)



def student_register(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")  # Change 'username' to 'email'
        password = request.POST.get("password")

        # Check if the email already exists in the database
        existing_student = Student_data.objects.filter(email=email).first()

        if existing_student:
            # If a student with the same email already exists, handle the error here
            error_message = "Email already exists. Please use a different email."
            return render(request, "Student_Login.html", {"error_message": error_message})

        # If the email is unique, create a new student record
        student = Student_data(
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=password
        )
        student.save()

        return redirect('Student_Login')

    return render(request, "Student_Login.html")


def faculty_login(request):
    if request.method == "POST":
        email = request.POST.get("email")  # Change 'username' to 'email'
        password = request.POST.get("password")

        try:
            faculty = Faculty_profile.objects.get(email=email, password=password)  # Change 'username' to 'email'
            # If the login is successful, store the email in the session
            request.session['email'] = email
            return redirect('faculty_dashboard')  # Redirect to faculty dashboard or the appropriate URL
        except Faculty_profile.DoesNotExist:
            messages.error(request, "Invalid email or password")
            return render(request, "Faculty_Login.html")

    return render(request, "Faculty_Login.html")



def faculty_dashboard(request):
    email = request.session.get('email')
    print(email)
    faculty_mappings = StudentFacultyMapping.objects.filter(faculty_email=email)
    print(faculty_mappings)
    faculty_name = ""
    if faculty_mappings.exists():
        faculty_name = faculty_mappings.first().faculty_name
    return render(request, 'faculty_dashboard.html', {"faculty_name": faculty_name, 'faculty_mappings': faculty_mappings, 'email': email})

def hod_login(request):
    if request.method == "POST":
        email = request.POST.get("username")
        password = request.POST.get("password")

        try:
            hod = HOD_profile.objects.get(email=email, password=password)
            return render(request, 'hod_dashboard.html')
        except HOD_profile.DoesNotExist:
            error_message = "Invalid username or password"
            messages.error(request, error_message)  # Add this line to set the error message
            return render(request, "HOD_Login.html")

    return render(request, "HOD_Login.html")
@login_required
def hod_dashboard(request):
    if request.method == 'POST':
        excel_file1 = request.FILES.get('excel_file1')
        excel_file2 = request.FILES.get('excel_file2')

        if excel_file1 and excel_file2:
            uploaded_files = UploadedFiles(file1=excel_file1, file2=excel_file2)
            uploaded_files.save()

    return render(request, 'hod_dashboard.html')


def generate_mapping_excel(request):
    if request.method == 'POST':
        faculty_file = request.FILES.get('faculty_file')
        student_file = request.FILES.get('student_file')

        if faculty_file and student_file:
            faculty_wb = openpyxl.load_workbook(faculty_file)
            faculty_sheet = faculty_wb.active

            student_wb = openpyxl.load_workbook(student_file)
            student_sheet = student_wb.active

            mapping = {'Professor': 15, 'Associate Professor': 15, 'Asst.Prof': 10, 'HOD': 15}
            faculty_mapping = {}

            faculty_rows = list(faculty_sheet.iter_rows(values_only=True))
            student_rows = list(student_sheet.iter_rows(values_only=True))

            faculty_rows.pop(0)  # Remove header row
            student_rows.pop(0)  # Remove header row

            for faculty_row in faculty_rows:
                designation = faculty_row[4]
                if designation in mapping:
                    num_students = mapping[designation]
                    assigned_students = student_rows[:num_students]
                    student_rows = student_rows[num_students:]  # Update student_rows for the next faculty member
                    faculty_mapping[faculty_row[1]] = {
                        'faculty_info': faculty_row,
                        'assigned_students': assigned_students
                    }

            mapped_wb = openpyxl.Workbook()
            mapped_sheet = mapped_wb.active
            mapped_sheet.append(
                ['S.No', 'Registration Number', 'Name', 'Branch/Specialization', 'Student Email', 'Faculty Name',
                 'Faculty Email', 'Faculty Phone'])

            s_no = 1

            for faculty_name, faculty_data in faculty_mapping.items():
                faculty_info = faculty_data['faculty_info']
                assigned_students = faculty_data['assigned_students']

                for student_row in assigned_students:
                    student_email = student_row[6]  # Assuming the student email is in the 7th column

                    # Check if student_email is not empty before creating the mapping entry
                    if student_email:
                        faculty_email = faculty_info[6]
                        faculty_phone = faculty_info[5]

                        mapped_sheet.append([
                            s_no,
                            student_row[1],
                            student_row[2],
                            student_row[3],
                            student_row[6],
                            faculty_name,
                            faculty_email,
                            faculty_phone
                        ])
                        s_no += 1

                        # Create a mapping entry and save to the database
                        mapping_entry = StudentFacultyMapping(
                            student_registration_number=student_row[1],
                            student_name=student_row[2],
                            student_branch=student_row[3],
                            student_email=student_email,
                            faculty_name=faculty_name,
                            faculty_email=faculty_email,
                            faculty_phone=faculty_phone
                        )
                        mapping_entry.save()

            excel_response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_response['Content-Disposition'] = 'attachment; filename=student_faculty_mapping.xlsx'
            mapped_wb.save(excel_response)

            return excel_response

    return render(request, 'hod_dashboard.html')


def download_mapped_excel(request):
    # Retrieve data from the StudentFacultyMapping model
    mapping_data = StudentFacultyMapping.objects.all()

    # Prepare Excel data
    mapped_wb = openpyxl.Workbook()
    mapped_sheet = mapped_wb.active
    mapped_sheet.append(
        ['S.No', 'Registration Number', 'Name', 'Branch/Specialization', 'Student Email', 'Faculty Name',
         'Faculty Email', 'Faculty Phone'])

    for index, entry in enumerate(mapping_data, start=1):
        mapped_sheet.append([
            index,
            entry.student_registration_number,
            entry.student_name,
            entry.student_branch,
            entry.student_email,
            entry.faculty_name,
            entry.faculty_email,
            entry.faculty_phone
        ])

    excel_response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    excel_response['Content-Disposition'] = 'attachment; filename=student_faculty_mapping.xlsx'
    mapped_wb.save(excel_response)

    return excel_response

def mapping_list(request):
    mappings = StudentFacultyMapping.objects.all()
    return render(request, 'mapping_list.html', {'mappings': mappings})



from django.shortcuts import render, redirect
from django.contrib import messages

def student_files(request):
    roll_number = None
    message = None

    if request.method == 'POST':
        roll_number = request.POST.get('roll_number')  # Get the roll number from the form

        # Check if the roll number already exists in the MyUploadedFiles model
        if MyUploadedFiles.objects.filter(roll_number=roll_number).exists():
            # If the roll number already exists, update the existing record
            uploaded_files = MyUploadedFiles.objects.get(roll_number=roll_number)

            # Update only the weekly report if provided in the form
            weekly_report = request.FILES.get('weekly_report')
            if weekly_report:
                uploaded_files.weekly_report = weekly_report
                uploaded_files.save()
                message = "Weekly report updated successfully!"

            # Update other documents if provided in the form
            certificate = request.FILES.get('certificate')
            ppt = request.FILES.get('ppt')
            final_project_report = request.FILES.get('final_project_report')

            if certificate:
                uploaded_files.certificate = certificate
            if ppt:
                uploaded_files.ppt = ppt
            if final_project_report:
                uploaded_files.final_project_report = final_project_report

            # Save the updated record for other documents if provided
            uploaded_files.save()

            if not message:
                message = "Documents updated successfully!"

        else:
            weekly_report = request.FILES.get('weekly_report')
            certificate = request.FILES.get('certificate')
            ppt = request.FILES.get('ppt')
            final_project_report = request.FILES.get('final_project_report')

            if roll_number and weekly_report:
                # Create a new record if the roll number does not exist
                uploaded_files = MyUploadedFiles(
                    roll_number=roll_number,
                    weekly_report=weekly_report,
                    certificate=certificate,
                    ppt=ppt,
                    final_project_report=final_project_report,
                    student=request.user,
                )
                uploaded_files.save()
                message = "Files uploaded successfully!"
            else:
                messages.error(request, "Please provide a roll number and weekly report")

    return render(request, 'student_files.html', {'roll_number': roll_number, 'message': message})


from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404
from .models import MyUploadedFiles
import os


def download_file(request, file_id):
    uploaded_file = get_object_or_404(MyUploadedFiles, id=file_id)
    file_path = uploaded_file.weekly_report.path  # Assuming this is the correct path
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = FileResponse(file)
        return response
    else:
        raise Http404("File not found")

from .models import MyUploadedFiles

def faculty_files_view(request):
    files = MyUploadedFiles.objects.all()
    return render(request, 'faculty_files_view.html', {'files': files})


def faculty_activities(request, roll_number):
    files = MyUploadedFiles.objects.filter(roll_number=roll_number)
    return render(request, 'faculty_activities.html', {'files': files})


def hod_activities(request,roll_number):
    files = MyUploadedFiles.objects.filter(roll_number=roll_number)
    return render(request, 'hod_activities.html', {'files': files})

from .models import MyUploadedFiles


def update_marks(request, file_id):
    if request.method == 'POST':
        file = get_object_or_404(MyUploadedFiles, pk=file_id)
        marks1 = float(request.POST.get('marks1', 0))  # Get marks1 as a float (default to 0 if not provided)
        marks2 = float(request.POST.get('marks2', 0))  # Get marks2 as a float (default to 0 if not provided)
        marks3 = float(request.POST.get('marks3', 0))  # Get marks3 as a float (default to 0 if not provided)
        marks4 = float(request.POST.get('marks4', 0))  # Get marks4 as a float (default to 0 if not provided)
        marks5 = float(request.POST.get('marks5', 0))  # Get marks5 as a float (default to 0 if not provided)
        marks6 = float(request.POST.get('marks6', 0))  # Get marks6 as a float (default to 0 if not provided)
        marks7 = float(request.POST.get('marks7', 0))  # Get marks7 as a float (default to 0 if not provided)
        marks8 = float(request.POST.get('marks8', 0))  # Get marks8 as a float (default to 0 if not provided)

        # Calculate the average marks
        total_marks = marks1 + marks2 + marks3 + marks4 + marks5 + marks6 + marks7 + marks8
        total_reviews = 8  # Assuming there are 8 reviews, update this number accordingly

        if total_reviews > 0:
            average_marks = total_marks / total_reviews
        else:
            average_marks = 0.0  # Avoid division by zero

        # Update the review fields and average marks in the file object
        file.marks1 = marks1
        file.marks2 = marks2
        file.marks3 = marks3
        file.marks4 = marks4
        file.marks5 = marks5
        file.marks6 = marks6
        file.marks7 = marks7
        file.marks8 = marks8
        file.average_marks = average_marks  # Create an average_marks field in your model

        file.save()

        return redirect('faculty_activities', roll_number=file.roll_number)
    else:
        pass


def student_marks_view(request):
    if request.method == 'POST':
        roll_number = request.POST.get('roll_number')
        marks = MyUploadedFiles.objects.filter(roll_number=roll_number)

        # Calculate and set the average marks for each record
        for mark in marks:
            total_marks = (
                mark.marks1 + mark.marks2 + mark.marks3 + mark.marks4 +
                mark.marks5 + mark.marks6 + mark.marks7 + mark.marks8
            )
            mark.average_marks = total_marks / 8  # Assuming 8 reviews

        return render(request, 'student_marks.html', {'marks': marks, 'roll_number': roll_number})
    else:
        return render(request, 'student_marks.html')

import openpyxl
from django.http import HttpResponse
from .models import MyUploadedFiles

def download_student_marks_excel(request):
    # Get the student data with the required fields
    student_data = MyUploadedFiles.objects.values('roll_number', 'marks1', 'marks2', 'marks3', 'marks4', 'marks5', 'marks6', 'marks7', 'marks8', 'average_marks')

    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Roll Number', 'Marks1', 'Marks2', 'Marks3', 'Marks4', 'Marks5', 'Marks6', 'Marks7', 'Marks8', 'Average Marks']
    worksheet.append(headers)

    # Add student data to the worksheet
    for data in student_data:
        row = [
            data['roll_number'],
            data['marks1'],
            data['marks2'],
            data['marks3'],
            data['marks4'],
            data['marks5'],
            data['marks6'],
            data['marks7'],
            data['marks8'],
            data['average_marks'],
        ]
        worksheet.append(row)

    # Create an HttpResponse with the Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_marks.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response
from django.http import JsonResponse

# Function to generate chatbot responses based on different message conditions
def get_chatbot_response(message):
    if message.lower() in ['hi', 'hello']:
        return "Hello! Welcome to the chatbot."

    elif message.lower() == 'how are you':
        return "I'm fine, thank you. How about you?"

    elif message.lower() == 'i love you':
        return "Sorry, I am already committed! 😄"

    else:
        return "I'm not sure how to respond to that."

def chatbot(request):
    if request.method == 'POST':
        message = request.POST.get('message')  # Assuming the message is sent via POST request
        response = get_chatbot_response(message)
        return JsonResponse({'response': response})
    else:
        return JsonResponse({'error': 'Invalid request method'})


from .models import MyUploadedFiles
def hod_marks_view(request):
    uploaded_files=MyUploadedFiles.objects.all()
    context={'uploaded_files': uploaded_files}
    return render(request,'hod_marks_view.html',context)


from .models import PrivateFile

from django.http import HttpResponse
from datetime import datetime

# def student_private_files(request):
#     if request.method == 'POST' and request.FILES.getlist('file'):
#         for uploaded_file in request.FILES.getlist('file'):
#             private_file = PrivateFile(file=uploaded_file)
#             private_file.save()
#
#     files = PrivateFile.objects.all()
#
#     return render(request, 'student_private_files.html', {'files': files})
def student_private_files(request, email):
    student_mapping = StudentFacultyMapping.objects.get(student_email=email)

    if request.method == 'POST' and request.FILES.getlist('file'):
        for uploaded_file in request.FILES.getlist('file'):
            private_file = PrivateFile(file=uploaded_file, student_mapping=student_mapping)
            private_file.save()

    files = PrivateFile.objects.filter(student_mapping=student_mapping)

    return render(request, 'student_private_files.html', {'files': files, 'email': email})



def download_file(request, file_id):
    file = get_object_or_404(PrivateFile, id=file_id)
    response = HttpResponse(file.file, content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{file.file.name}"'
    return response


from .models import Company
import pandas as pd

# def upload_file(request):
#     if request.method == 'POST':
#         file = request.FILES['file']
#         if file.name.endswith('.csv') or file.name.endswith('.xlsx'):
#             df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)
#
#             for index, row in df.iterrows():
#                 Company.objects.create(
#                     internship_dates=row['internship_dates'],
#                     company_name=row['company_name'],
#                     job_location=row['job_location'],
#                     role=row['role'],
#                     eligibility_criteria=row['eligibility_criteria'],
#                     stipend=row['stipend']
#                 )
#
#             return redirect('admin_dashboard')  # Redirect to the admin dashboard after successful upload
#
#     return render(request, 'admin_dashboard.html')
import pandas as pd
from django.shortcuts import render, redirect
from .models import Company

def upload_file(request):
    if request.method == 'POST':
        file = request.FILES['file']
        if file.name.endswith('.csv') or file.name.endswith('.xlsx'):
            df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)

            for index, row in df.iterrows():
                company_name = row['company_name']
                application_url = row['application_url']

                # Check if the company with the same name already exists
                existing_company = Company.objects.filter(company_name=company_name).first()

                if existing_company is None:
                    # Company doesn't exist, create a new entry
                    Company.objects.create(
                        internship_dates=row['internship_dates'],
                        company_name=company_name,
                        job_location=row['job_location'],
                        role=row['role'],
                        eligibility_criteria=row['eligibility_criteria'],
                        stipend=row['stipend'],
                        application_url=application_url if pd.notna(application_url) else ''  # Set application_url to '' if it's NaN
                    )

            return redirect('admin_dashboard')  # Redirect to the admin dashboard after successful upload

    return render(request, 'admin_dashboard.html')

def admin_dashboard(request):
    # your view logic here
    return render(request, 'admin_dashboard.html')


from .models import Company

def internship_opportunities(request):
    companies = Company.objects.all()
    return render(request, 'internship.html', {'companies': companies})
def resume_builder(request):
    return render(request, 'resume_builder.html')



import speech_recognition as sr
import pyttsx3
from django.urls import reverse

def speech_recognition_and_execution(request):
    def say(text):
        engine = pyttsx3.init()
        engine.say(text)
        engine.runAndWait()

    def take_command():
        r = sr.Recognizer()
        with sr.Microphone() as source:
            r.pause_threshold = 0.6
            audio = r.listen(source)
            try:
                print("recognizing.........")
                query = r.recognize_google(audio, language="en-in")
                print(f"user said: {query}")
                return query.lower()
            except Exception as e:
                return "some Error has been occurred please wait for some time"

    say("hello, I am IMS A I created by bolgum hemanth goud")

    while True:
        print("listening........")
        query = take_command()

        if "upload files" in query:
            say("Opening upload files, sir...")
            return redirect(reverse('student_files'))

        elif "open resume" in query:
            say("Opening your updated resume, sir...")
            return redirect(reverse('resume_builder'))

        elif "check marks" in query:
            say("Please enter your roll number to check your marks, sir...")
            return redirect(reverse('student_marks_view'))

        elif "resource center" in query:
            say("Welcome to Resource Center sir...")
            return redirect(reverse('resource_center/'))


    # Render a response if needed
    return render(request, 'student_dashboard.html')



# from .models import Document
#
# def resource_center(request):
#     # Fetch all documents
#     documents = Document.objects.all()
#
#     # Organize documents by domain
#     organized_documents = {}
#     for document in documents:
#         if document.domain not in organized_documents:
#             organized_documents[document.domain] = []
#         organized_documents[document.domain].append(document)
#
#     return render(request, 'resource_center.html', {'organized_documents': organized_documents})

from django.shortcuts import render, redirect
from .models import Document

def resource_center(request):
    if request.method == 'POST':
        # Handle the case where a document link is clicked
        document_id = request.POST.get('document_id')
        document = Document.objects.get(pk=document_id)
        return redirect(document.link)

    # Fetch all documents
    documents = Document.objects.all()

    # Organize documents by domain
    organized_documents = {}
    for document in documents:
        if document.domain not in organized_documents:
            organized_documents[document.domain] = []
        organized_documents[document.domain].append(document)

    return render(request, 'resource_center.html', {'organized_documents': organized_documents})


from .models import Circular

from django.contrib import messages  # Import messages module for displaying messages



def circular(request):
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        # Check if subject and message are not empty
        if subject and message:
            Circular.objects.create(subject=subject, message=message)
            messages.success(request, 'Circular published successfully.')
        else:
            messages.error(request, 'Subject and message cannot be empty.')

    return render(request, 'hod_dashboard.html')

from django.shortcuts import render
from .models import MyUploadedFiles

def student_performance(request):
    if request.method == 'POST':
        roll_number = request.POST.get('roll_number')
        student_files = MyUploadedFiles.objects.filter(roll_number=roll_number).first()
        if student_files:
            marks = [
                student_files.marks1,
                student_files.marks2,
                student_files.marks3,
                student_files.marks4,
                student_files.marks5,
                student_files.marks6,
                student_files.marks7,
                student_files.marks8
            ]
            return render(request, 'performance.html', {'marks': marks})
        else:
            return render(request, 'performance.html', {'error_message': 'Student not found.'})
    else:
        return render(request, 'performance.html')
